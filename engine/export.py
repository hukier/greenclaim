"""Export del borrador de reconciliación a Excel (paso 6 — entregable final).

Toma la tabla final completa (núcleo + capa derivada) y produce un ``.xlsx``
estructurado, revisable por el equipo ambiental, con dos hojas:

  - ``Reconciliacion``: todas las filas, con color de fondo por estado.
  - ``Resumen``: totales por energético, conteo por estado, metadata de la
    planta y una advertencia visible si hay desviaciones sin justificar.

PDF: deseable post-MVP (una página con la misma info). NO se implementa aquí.

Headless: no depende de la UI. Reglas de negocio (estados, justificación) ya
vienen resueltas en pasos anteriores; aquí solo se presentan.
"""
from __future__ import annotations

from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Orden EXACTO de columnas en la hoja de reconciliación.
COLUMNAS_REPORTE = [
    "periodo", "energetico", "tipo", "unidad",
    "valor_op", "valor_cont", "brecha_abs", "brecha_pct",
    "energia_gj", "emisiones_tco2eq", "es_biogenico",
    "estado", "justificacion",
]

META_REQUERIDA = ("planta", "periodo", "generado_en")

# Colores de fondo por estado (tonos claros).
_FILL_ESTADO = {
    "cuadrado": PatternFill("solid", fgColor="C6EFCE"),    # verde claro
    "desviacion": PatternFill("solid", fgColor="FFC7CE"),  # rojo claro
    "pendiente": PatternFill("solid", fgColor="FFEB9C"),   # amarillo claro
}
_FILL_ADVERTENCIA = PatternFill("solid", fgColor="FFC7CE")
_NEGRITA = Font(bold=True)


def _py(valor):
    """Convierte tipos de numpy/pandas a nativos para openpyxl; NaN → None."""
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(valor, "item"):   # numpy scalar
        return valor.item()
    return valor


def _sin_justificar(df: pd.DataFrame) -> pd.Series:
    """Filas en desviación sin justificación (NaN o texto vacío)."""
    just = df["justificacion"].astype("object")
    vacia = just.isna() | just.map(lambda v: str(v).strip() == "")
    return (df["estado"] == "desviacion") & vacia


def _preparar(df_final: pd.DataFrame) -> pd.DataFrame:
    """Garantiza todas las columnas del reporte en orden (rellena ausentes)."""
    df = df_final.copy()
    if "justificacion" not in df.columns:
        df["justificacion"] = ""
    for col in COLUMNAS_REPORTE:
        if col not in df.columns:
            df[col] = pd.NA
    return df[COLUMNAS_REPORTE]


def _resumen_datos(df: pd.DataFrame) -> dict:
    """Calcula los agregados del resumen a partir de la tabla del reporte."""
    por_energetico = (
        df.groupby("energetico")[["energia_gj", "emisiones_tco2eq"]]
        .sum(min_count=1)
        .sort_index()
    )
    por_estado = df["estado"].value_counts().to_dict()
    n_sin_just = int(_sin_justificar(df).sum())
    return {
        "n_filas": len(df),
        "por_estado": por_estado,
        "totales_por_energetico": {
            e: {"energia_gj": _py(fila["energia_gj"]),
                "emisiones_tco2eq": _py(fila["emisiones_tco2eq"])}
            for e, fila in por_energetico.iterrows()
        },
        "energia_gj_total": _py(df["energia_gj"].sum(min_count=1)),
        "emisiones_tco2eq_total": _py(df["emisiones_tco2eq"].sum(min_count=1)),
        "desviaciones_sin_justificar": n_sin_just,
    }


def _hoja_reconciliacion(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Reconciliacion"
    ws.append(COLUMNAS_REPORTE)
    for celda in ws[1]:
        celda.font = _NEGRITA
    ws.freeze_panes = "A2"

    for _, fila in df.iterrows():
        ws.append([_py(fila[c]) for c in COLUMNAS_REPORTE])
        relleno = _FILL_ESTADO.get(fila["estado"])
        if relleno is not None:
            for celda in ws[ws.max_row]:
                celda.fill = relleno


def _hoja_resumen(wb: Workbook, df: pd.DataFrame, meta: dict, datos: dict) -> None:
    ws = wb.create_sheet("Resumen")
    r = 1

    def escribir(texto, valor=None, *, negrita=False, relleno=None):
        nonlocal r
        c1 = ws.cell(row=r, column=1, value=texto)
        if negrita:
            c1.font = _NEGRITA
        if relleno is not None:
            c1.fill = relleno
        if valor is not None:
            ws.cell(row=r, column=2, value=_py(valor))
        r += 1

    escribir("Borrador de reconciliación — GreenClaim", negrita=True)
    r += 1
    escribir("Planta:", meta["planta"])
    escribir("Período:", meta["periodo"])
    escribir("Generado:", meta["generado_en"])
    r += 1

    n_sin = datos["desviaciones_sin_justificar"]
    if n_sin:
        escribir(f"⚠ {n_sin} desviación(es) sin justificar — requieren nota para cerrar.",
                 negrita=True, relleno=_FILL_ADVERTENCIA)
    else:
        escribir("Sin desviaciones sin justificar.")
    r += 1

    escribir("Totales por energético (emisiones = solo fósiles)", negrita=True)
    for celda, val in zip(("A", "B", "C"), ("energetico", "energia_gj", "emisiones_tco2eq")):
        c = ws[f"{celda}{r}"]
        c.value = val
        c.font = _NEGRITA
    r += 1
    for energetico, tot in datos["totales_por_energetico"].items():
        ws.cell(row=r, column=1, value=energetico)
        ws.cell(row=r, column=2, value=tot["energia_gj"])
        ws.cell(row=r, column=3, value=tot["emisiones_tco2eq"])
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = _NEGRITA
    ws.cell(row=r, column=2, value=datos["energia_gj_total"])
    ws.cell(row=r, column=3, value=datos["emisiones_tco2eq_total"])
    r += 2

    escribir("Conteo por estado", negrita=True)
    for estado in ("cuadrado", "desviacion", "pendiente"):
        if estado in datos["por_estado"]:
            escribir(estado, datos["por_estado"][estado])

    ws.column_dimensions["A"].width = 42
    for col in ("B", "C"):
        ws.column_dimensions[col].width = 18


def exportar_excel(df_final: pd.DataFrame, meta: dict, path: str) -> dict:
    """Escribe el borrador ``.xlsx`` con las hojas Reconciliacion y Resumen.

    ``meta`` debe traer al menos ``planta``, ``periodo`` y ``generado_en``.
    Devuelve un dict con el resumen de lo exportado (totales, conteos por estado
    y nº de desviaciones sin justificar) — útil para la consola y los tests.
    """
    faltan = [k for k in META_REQUERIDA if k not in meta]
    if faltan:
        raise ValueError("Falta(n) clave(s) en 'meta': " + ", ".join(faltan))

    df = _preparar(df_final)
    datos = _resumen_datos(df)

    wb = Workbook()
    _hoja_reconciliacion(wb, df)
    _hoja_resumen(wb, df, meta, datos)
    wb.save(path)

    datos["path"] = path
    return datos
