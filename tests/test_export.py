"""Pruebas del export del borrador a Excel (paso 6)."""
import os
import sys

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from engine import export  # noqa: E402
from engine.export import exportar_excel, COLUMNAS_REPORTE  # noqa: E402

META = {"planta": "Planta Tipo", "periodo": "2025-06", "generado_en": "2026-06-14"}


def _final(filas):
    """Tabla final (paso 5) con las columnas que produce el motor."""
    cols = ["periodo", "energetico", "tipo", "unidad", "valor_op", "valor_cont",
            "brecha_abs", "brecha_pct", "estado", "energia_gj",
            "emisiones_tco2eq", "es_biogenico"]
    return pd.DataFrame(filas, columns=cols)


@pytest.fixture
def df_ok():
    return _final([
        ("2025-06", "gas_natural", "combustible_fosil", "Nm3", 1000, 980,
         20, 2.04, "cuadrado", 39.34, 2.21, False),
        ("2025-06", "electricidad_comprada", "electricidad_red", "kWh", 1000, 900,
         100, 11.1, "desviacion", 3.6, 0.43, False),
        ("2025-06", "biomasa", "combustible_biogenico", "ton", 50, None,
         None, None, "pendiente", 750.0, 0.0, True),
    ])


def test_dos_hojas_exactas(tmp_path, df_ok):
    path = str(tmp_path / "b.xlsx")
    exportar_excel(df_ok, META, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Reconciliacion", "Resumen"]


def test_orden_de_columnas_reconciliacion(tmp_path, df_ok):
    path = str(tmp_path / "b.xlsx")
    exportar_excel(df_ok, META, path)
    ws = load_workbook(path)["Reconciliacion"]
    encabezado = [c.value for c in ws[1]]
    assert encabezado == COLUMNAS_REPORTE


def test_totales_cuadran_con_entrada(tmp_path, df_ok):
    path = str(tmp_path / "b.xlsx")
    datos = exportar_excel(df_ok, META, path)
    assert datos["energia_gj_total"] == pytest.approx(39.34 + 3.6 + 750.0)
    assert datos["emisiones_tco2eq_total"] == pytest.approx(2.21 + 0.43 + 0.0)
    assert datos["totales_por_energetico"]["gas_natural"]["energia_gj"] == pytest.approx(39.34)
    assert datos["por_estado"] == {"cuadrado": 1, "desviacion": 1, "pendiente": 1}


def test_advertencia_si_desviacion_sin_justificar(tmp_path, df_ok):
    path = str(tmp_path / "b.xlsx")
    datos = exportar_excel(df_ok, META, path)
    assert datos["desviaciones_sin_justificar"] == 1
    # La advertencia debe estar visible en la hoja Resumen.
    ws = load_workbook(path)["Resumen"]
    textos = [str(c.value) for fila in ws.iter_rows() for c in fila if c.value]
    assert any("sin justificar" in t for t in textos)


def test_sin_advertencia_si_justificada(tmp_path, df_ok):
    df = df_ok.copy()
    df["justificacion"] = ["", "parada programada 12/06", ""]
    path = str(tmp_path / "b.xlsx")
    datos = exportar_excel(df, META, path)
    assert datos["desviaciones_sin_justificar"] == 0


def test_archivo_relegible_con_openpyxl(tmp_path, df_ok):
    path = str(tmp_path / "b.xlsx")
    exportar_excel(df_ok, META, path)
    wb = load_workbook(path)   # no debe lanzar
    assert wb["Reconciliacion"].max_row == 1 + len(df_ok)


def test_columna_justificacion_se_agrega_si_falta(tmp_path, df_ok):
    assert "justificacion" not in df_ok.columns
    path = str(tmp_path / "b.xlsx")
    exportar_excel(df_ok, META, path)
    ws = load_workbook(path)["Reconciliacion"]
    assert [c.value for c in ws[1]][-1] == "justificacion"


def test_meta_incompleta_falla(tmp_path, df_ok):
    with pytest.raises(ValueError, match="meta"):
        exportar_excel(df_ok, {"planta": "X"}, str(tmp_path / "b.xlsx"))
